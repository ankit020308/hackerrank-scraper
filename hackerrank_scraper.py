#!/usr/bin/env python3
"""
HackerRank Solved Problems Scraper
Fetches accepted submissions, enriches each with difficulty via the
challenge API (difficulty isn't in the submission list), then exports
Medium / Hard / Expert / Advanced results to Excel.

HOW TO GET YOUR COOKIE STRING (one-time):
  1. Log in to hackerrank.com in Chrome or Safari
  2. DevTools → Network tab → refresh the page
  3. Click any request to hackerrank.com
  4. Scroll to Request Headers → find "cookie:"
  5. Copy the entire value and paste when prompted
"""

import sys
import time
import warnings
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)

warnings.filterwarnings("ignore")

BASE_URL = "https://www.hackerrank.com"
API      = f"{BASE_URL}/rest/contests/master"

TARGET   = {"medium", "hard", "expert", "advanced"}
LABEL    = {"easy": "Easy", "medium": "Medium", "hard": "Hard",
            "expert": "Expert", "advanced": "Advanced"}
ORDER    = {"Medium": 0, "Hard": 1, "Expert": 2, "Advanced": 3}
COLORS   = {
    "Medium":   ("FFF3CD", "856404"),
    "Hard":     ("F8D7DA", "721C24"),
    "Expert":   ("E8D5F5", "6A0DAD"),
    "Advanced": ("D1ECF1", "0C5460"),
}


# ── session ──────────────────────────────────────────────────────────────────

def build_session(cookie_string: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent":       ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                             "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                             "Version/26.3.1 Safari/605.1.15"),
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          f"{BASE_URL}/submissions/all",
        "Cookie":           cookie_string,
    })
    return s


def get_username(session: requests.Session):
    r = session.get(f"{BASE_URL}/rest/hackers/me", timeout=10)
    if r.status_code == 200:
        return r.json().get("model", {}).get("username")
    return None


# ── data fetching ─────────────────────────────────────────────────────────────

def fetch_all_accepted(session: requests.Session) -> dict:
    """Return {slug: submission} for every unique accepted problem."""
    solved, offset, limit = {}, 0, 20
    total = None

    print("Fetching submissions…")
    while True:
        r = session.get(
            f"{API}/submissions",
            params={"offset": offset, "limit": limit},
            timeout=15,
        )
        if r.status_code != 200:
            print(f"\n  HTTP {r.status_code} — stopping.")
            break

        data   = r.json()
        models = data.get("models", [])

        if total is None:
            total = data.get("total", "?")
            print(f"  Total submissions on record: {total}")

        if not models:
            break

        for sub in models:
            if sub.get("status") != "Accepted":
                continue
            slug = (sub.get("challenge") or {}).get("slug")
            if slug and slug not in solved:
                solved[slug] = sub

        offset += limit
        shown = min(offset, int(total)) if str(total).isdigit() else offset
        print(f"  Scanned {shown}/{total} …", end="\r")
        time.sleep(0.25)

        if len(models) < limit:
            break

    print(f"\n  Unique accepted problems: {len(solved)}")
    return solved


def fetch_challenge(session: requests.Session, slug: str) -> dict:
    """Return the challenge model dict, or {} on failure."""
    r = session.get(f"{API}/challenges/{slug}", timeout=10)
    if r.status_code == 200:
        return r.json().get("model", {})
    return {}


def enrich_and_filter(session: requests.Session, solved_map: dict) -> list:
    """
    For each accepted slug, hit the challenge API to get difficulty and track.
    Keep only TARGET difficulties.
    """
    print("\nFetching difficulty for each problem…")
    problems = []
    slugs    = list(solved_map.keys())

    for idx, slug in enumerate(slugs, 1):
        sub    = solved_map[slug]
        detail = fetch_challenge(session, slug)

        diff_raw = (detail.get("difficulty_name") or "").strip().lower()

        if diff_raw not in TARGET:
            label = diff_raw or "unknown"
            print(f"  [{idx}/{len(slugs)}] skip  {slug}  ({label})      ", end="\r")
            time.sleep(0.15)
            continue

        track      = detail.get("track") or {}
        name       = detail.get("name") or sub.get("challenge", {}).get("name") or slug
        domain     = track.get("track_name") or track.get("name") or "N/A"
        score      = sub.get("score", "N/A")
        ts         = sub.get("created_at")
        solved_at  = (datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d")
                      if isinstance(ts, (int, float)) else str(ts or "")[:10]) or "N/A"
        url        = f"{BASE_URL}/challenges/{slug}/problem"

        problems.append({
            "name":       name,
            "difficulty": LABEL[diff_raw],
            "domain":     domain,
            "score":      score,
            "solved_at":  solved_at,
            "url":        url,
        })
        print(f"  [{idx}/{len(slugs)}] + {name}  ({LABEL[diff_raw]})          ", end="\r")
        time.sleep(0.2)

    print(f"\n  Kept {len(problems)} problems (Easy excluded)")
    return problems


# ── excel ─────────────────────────────────────────────────────────────────────

def build_excel(problems: list, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solved Problems"

    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", start_color="1A1A2E")
    alt_fill = PatternFill("solid", start_color="F8F9FA")
    dom_fill = PatternFill("solid", start_color="EBF5FB")
    wht_fill = PatternFill("solid", start_color="FFFFFF")

    # title row
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "HackerRank — Solved Problems (Medium → Hard → Expert → Advanced)"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="0A3D62")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # header row
    headers    = ["#", "Problem Name", "Difficulty", "Domain", "Score", "Solved At", "Link"]
    col_widths = [5,   40,             12,           22,       8,       13,           9]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=2, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28

    # data rows
    for i, p in enumerate(problems, 1):
        row = i + 2
        diff_bg, diff_fg = COLORS.get(p["difficulty"], ("FFFFFF", "000000"))
        row_fill = alt_fill if i % 2 == 0 else wht_fill

        vals   = [i, p["name"], p["difficulty"], p["domain"],
                  p["score"], p["solved_at"], p["url"]]
        aligns = [center, left, center, left, center, center, left]

        for ci, (val, align) in enumerate(zip(vals, aligns), 1):
            cell           = ws.cell(row=row, column=ci, value=val)
            cell.alignment = align
            cell.border    = border
            cell.font      = Font(name="Arial", size=10)

            if ci == 3:
                cell.fill = PatternFill("solid", start_color=diff_bg)
                cell.font = Font(name="Arial", size=10, bold=True, color=diff_fg)
            elif ci == 4:
                cell.fill = dom_fill
            else:
                cell.fill = row_fill

            if ci == 7 and val:
                cell.hyperlink = val
                cell.value     = "Open →"
                cell.font      = Font(name="Arial", size=10,
                                      color="0563C1", underline="single")

        ws.row_dimensions[row].height = 22

    # summary sheet
    ws2             = wb.create_sheet("Summary")
    ws2["A1"].value = "Summary"
    ws2["A1"].font  = Font(name="Arial", bold=True, size=14, color="0A3D62")
    ws2.row_dimensions[1].height = 30

    for ci, h in enumerate(["Metric", "Count"], 1):
        c           = ws2.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1A1A2E")
        c.alignment = center
        c.border    = border

    summary = [
        ("Total (Easy excluded)", len(problems)),
        ("Medium",   sum(1 for p in problems if p["difficulty"] == "Medium")),
        ("Hard",     sum(1 for p in problems if p["difficulty"] == "Hard")),
        ("Expert",   sum(1 for p in problems if p["difficulty"] == "Expert")),
        ("Advanced", sum(1 for p in problems if p["difficulty"] == "Advanced")),
    ]
    for ri, (metric, val) in enumerate(summary, 3):
        ws2.cell(row=ri, column=1, value=metric).border = border
        ws2.cell(row=ri, column=2, value=val).border    = border
        ws2.cell(row=ri, column=1).font = Font(name="Arial", size=10)
        c2 = ws2.cell(row=ri, column=2)
        c2.font      = Font(name="Arial", size=10, bold=True)
        c2.alignment = center
        ws2.row_dimensions[ri].height = 20

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 10

    ws.freeze_panes = "A3"

    # print / PDF settings
    last_row = len(problems) + 2
    ws.print_area             = f"A1:G{last_row}"
    ws.print_title_rows       = "1:2"          # repeat title + header on every page
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0             # unlimited pages tall
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left      = 0.5
    ws.page_margins.right     = 0.5
    ws.page_margins.top       = 0.6
    ws.page_margins.bottom    = 0.6
    ws.page_margins.header    = 0.3
    ws.page_margins.footer    = 0.3
    ws.oddHeader.center.text  = "HackerRank — Solved Problems"
    ws.oddFooter.center.text  = "Page &P of &N"

    # same for summary sheet
    ws2.print_area             = "A1:B7"
    ws2.page_setup.orientation = "portrait"
    ws2.page_setup.paperSize   = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToPage   = True
    ws2.page_setup.fitToWidth  = 1
    ws2.page_setup.fitToHeight = 1

    wb.save(path)
    print(f"Saved: {path}")


# ── pdf ───────────────────────────────────────────────────────────────────────

DIFF_HEX = {
    "Medium":   (colors.HexColor("#856404"), colors.HexColor("#FFF3CD")),
    "Hard":     (colors.HexColor("#721C24"), colors.HexColor("#F8D7DA")),
    "Expert":   (colors.HexColor("#6A0DAD"), colors.HexColor("#E8D5F5")),
    "Advanced": (colors.HexColor("#0C5460"), colors.HexColor("#D1ECF1")),
}

def build_pdf(problems: list, path: str) -> None:
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title="HackerRank Solved Problems",
    )

    title_st = ParagraphStyle("title", fontSize=14, fontName="Helvetica-Bold",
                               textColor=colors.HexColor("#0A3D62"),
                               spaceAfter=6)
    cell_st  = ParagraphStyle("cell",  fontSize=8,  fontName="Helvetica",
                               leading=10)
    hdr_st   = ParagraphStyle("hdr",   fontSize=9,  fontName="Helvetica-Bold",
                               textColor=colors.white)

    headers = ["#", "Problem Name", "Difficulty", "Domain", "Score", "Solved At"]
    col_w   = [10*mm, 90*mm, 24*mm, 48*mm, 16*mm, 24*mm]

    rows = [[Paragraph(h, hdr_st) for h in headers]]
    for i, p in enumerate(problems, 1):
        rows.append([
            Paragraph(str(i),          cell_st),
            Paragraph(p["name"],       cell_st),
            Paragraph(p["difficulty"], cell_st),
            Paragraph(p["domain"],     cell_st),
            Paragraph(str(p["score"]), cell_st),
            Paragraph(p["solved_at"],  cell_st),
        ])

    # per-row difficulty colouring
    row_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#FFFFFF"), colors.HexColor("#F8F9FA")]),
        ("GRID",    (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",  (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]
    for i, p in enumerate(problems, 1):
        fg, bg = DIFF_HEX.get(p["difficulty"], (colors.black, colors.white))
        row_styles += [
            ("BACKGROUND", (2, i), (2, i), bg),
            ("TEXTCOLOR",  (2, i), (2, i), fg),
            ("FONTNAME",   (2, i), (2, i), "Helvetica-Bold"),
        ]

    table = Table(rows, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle(row_styles))

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(
            landscape(A4)[0] / 2,
            8 * mm,
            f"Page {doc.page}  —  HackerRank Solved Problems",
        )
        canvas.restoreState()

    story = [
        Paragraph("HackerRank — Solved Problems", title_st),
        Spacer(1, 4 * mm),
        table,
    ]
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    print(f"Saved: {path}")


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 60)
    print("  HackerRank Scraper — Cookie Auth")
    print("=" * 60)
    print()
    print("Paste your cookie string")
    print("(DevTools → Network → any hackerrank.com request → Request Headers → cookie:)")
    print()

    cookie_str = input("> ").strip()
    if not cookie_str:
        print("No cookie provided. Exiting.")
        sys.exit(1)

    session  = build_session(cookie_str)
    username = get_username(session)

    if not username:
        print("\nCould not verify login — cookie may be expired or wrong.")
        sys.exit(1)

    print(f"\nLogged in as: {username}\n")

    solved_map = fetch_all_accepted(session)
    if not solved_map:
        print("No accepted submissions found.")
        sys.exit(0)

    problems = enrich_and_filter(session, solved_map)
    if not problems:
        print("No Medium / Hard / Expert / Advanced problems found.")
        sys.exit(0)

    problems.sort(key=lambda p: (ORDER.get(p["difficulty"], 99), p["name"]))

    xlsx_out = "hackerrank_solved.xlsx"
    pdf_out  = "hackerrank_solved.pdf"
    build_excel(problems, xlsx_out)
    build_pdf(problems, pdf_out)

    print(f"\n{'='*60}")
    print(f"  Done!  {len(problems)} problems exported")
    for diff in ("Medium", "Hard", "Expert", "Advanced"):
        n = sum(1 for p in problems if p["difficulty"] == diff)
        if n:
            print(f"  {diff:<10} {n}")
    print(f"{'='*60}")
    print(f"\nExcel: open {xlsx_out}")
    print(f"PDF:   open {pdf_out}")


if __name__ == "__main__":
    main()
